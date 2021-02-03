import numpy as np 
import pandas as pd 
import openpyxl as op 
from openpyxl.styles import fills,colors,NamedStyle,Font,Side,Border,PatternFill,Alignment,Protection
import os
#后吧酒水出入统计,对接后吧

    

    

def beauty():
    wb = op.load_workbook("后吧数据统计.xlsx")
    ft = op.styles.Font(name='宋体', size=12, bold=False)  
    ft1 =  op.styles.Font(name='黑体', size=13, bold=False)  
    align = op.styles.Alignment(horizontal='center',vertical='center' )
    border = op.styles.Border(left=Side(border_style="thin"),
                                right=Side(border_style="thin"),         
                                top=Side(border_style="thin"),
                                bottom=Side(border_style="thin"))
    
    for sheet in wb:
        autofit(sheet)
        for row in sheet.rows:
            for cell in row:
                cell.font= ft
                cell.alignment =align
                cell.border = border
                if cell.row == 1 or cell.column ==1:
                    cell.font = ft1
    
    wb.save(DIR_DESKTOP+"后吧数据统计.xlsx")


def autofit(ws):
#使用：for循环遍历得出每列长度后形成字典数据来自动设置每列列宽。
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:

                len_cell = max([(len(line.encode('utf-8'))-len(line))/2+len(line) for line in str(cell.value).split('\n')])
                #dims[chr(64+cell.column)] = max((dims.get(chr(64+cell.column), 0), len(str(cell.value))))
                dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len_cell)
    for col, value in dims.items():

        ws.column_dimensions[col].width = value+2 if value+2<=50 else 50


def beauty(wb_name):
    wb = op.load_workbook(wb_name)
    ft = op.styles.Font(name='宋体', size=12, bold=False)  
    ft1 =  op.styles.Font(name='黑体', size=13, bold=False)  
    align = op.styles.Alignment(horizontal='center',vertical='center' )
    border = op.styles.Border(left=Side(border_style="thin"),
                                right=Side(border_style="thin"),         
                                top=Side(border_style="thin"),
                                bottom=Side(border_style="thin"))
    
    for sheet in wb:
        autofit(sheet)
        for row in sheet.rows:
            for cell in row:
                cell.font= ft
                cell.alignment =align
                cell.border = border
                if cell.row<3 or cell.column<3:
                    cell.font = ft1
    
    wb.save(wb_name)


def autofit(ws):

    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                len_cell = max([(len(line.encode('utf-8'))-len(line))/2+len(line) for line in str(cell.value).split('\n')])
                #dims[chr(64+cell.column)] = max((dims.get(chr(64+cell.column), 0), len(str(cell.value))))
                dims[cell.column_letter] = max(dims.get(cell.column_letter, 0), len_cell)
    for col, value in dims.items():

        ws.column_dimensions[col].width = value+2 if value+2<=50 else 50


def get_donate_table(dir):
    return pd.read_excel(dir + "赠送查询.xlsx")

def get_access_table(dir):
    return pd.read_excel(dir + "存取酒汇总.xlsx")

def get_detail_table(dir):
    return pd.read_excel(dir + "落单明细.xlsx")

def get_wine_type(dir_desktop):
    return pd.read_excel(dir_desktop+ "基础数据.xlsx", sheet_name='酒水类别')

def delete(dir):
    os.remove(dir + "存取酒汇总.xlsx")
    os.remove(dir + "落单明细.xlsx")


if __name__=="__main__":


    dir_desktop = os.path.join(os.path.expanduser("~"), 'Desktop').replace("\\", "/")+"/"
    dir = dir_desktop
    h_writer = pd.ExcelWriter(dir_desktop+"后吧数据统计.xlsx")
    # c_writer = pd.ExcelWriter(dir_desktop+"传送数据统计.xlsx")


    df_access = get_access_table(dir)
    df_detail = get_detail_table(dir)
    
    wine_type = get_wine_type(dir_desktop)

    df_detail = pd.merge(df_detail, wine_type, on="酒水项目", how='left')

    # !后吧:赠送,售出,存取
    h_access = df_access[['序号', '酒水类别', '酒水项目', '存酒数', '取酒数']]
    h_donate = pd.pivot_table(df_detail.query('类型=="经理赠送" & kind =="单品"'), index=['落单人','项目'], values=['数量', '金额'], aggfunc={'数量':np.sum, '金额':np.sum}, fill_value=0)

    h_detail = pd.pivot_table(df_detail.query('kind =="单品"'), index=['类型','类别','项目'], values=['数量', '金额'],  aggfunc={'数量':np.sum, '金额':np.sum}, fill_value=0)

    h_access.to_excel(h_writer,sheet_name='存取酒', index=False)
    
    h_donate.to_excel(h_writer,sheet_name='赠送')

    h_detail.to_excel(h_writer,sheet_name='售卖')

    h_writer.save()
    h_writer.close()
    delete(dir)
    # # !后吧:赠送,售出,存取

    # c_donate = pd.pivot_table(df_donate.query('类型 =="单品" and (传送 =="是")'), index=['赠送项目'], values='数量', aggfunc=np.sum, fill_value=0)

    # c_detail = pd.pivot_table(df_detail.query('类型 =="单品" and (传送 =="是")'), index=['类别','酒水项目'],values=['数量', '金额'],  aggfunc={'数量':np.sum, '金额':np.sum}, fill_value=0)

    # c_donate.to_excel(c_writer,sheet_name='赠送')

    # c_detail.to_excel(c_writer,sheet_name='售卖')
    
    # c_writer.save()
    # c_writer.close()



    beauty(dir_desktop+"后吧数据统计.xlsx")